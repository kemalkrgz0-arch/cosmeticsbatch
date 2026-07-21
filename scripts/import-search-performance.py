#!/usr/bin/env python3
"""Import an immutable XLSX search-performance export as reviewable TSV data."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as error:  # pragma: no cover - environment guidance
    raise SystemExit("openpyxl is required: python3 -m pip install openpyxl") from error


ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = ROOT / "data" / "search-performance"
SAFE_ID = re.compile(r"^[A-Z0-9][A-Z0-9-]{2,79}$")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[^\w.-]+", "-", name.strip(), flags=re.UNICODE).strip("-.")
    return cleaned or "sheet"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-id", required=True)
    parser.add_argument("--input", required=True, type=Path)
    args = parser.parse_args()

    if not SAFE_ID.fullmatch(args.source_id):
        raise SystemExit("source ID must contain only uppercase letters, digits and hyphens")
    source = args.input.expanduser().resolve()
    if source.suffix.lower() != ".xlsx" or not source.is_file():
        raise SystemExit(f"XLSX source not found: {source}")

    raw_dir = DATA_ROOT / "raw"
    output_dir = DATA_ROOT / "normalized" / args.source_id
    raw_target = raw_dir / f"{args.source_id}.xlsx"
    if raw_target.exists() or output_dir.exists():
        raise SystemExit(f"source ID already imported: {args.source_id}")

    raw_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True)
    shutil.copyfile(source, raw_target)

    workbook = load_workbook(raw_target, read_only=True, data_only=True)
    sheets = []
    used_names: set[str] = set()
    for index, worksheet in enumerate(workbook.worksheets, start=1):
        # Yandex Webmaster writes a <dimension> of A1:A1 on sheets that actually
        # carry hundreds of rows. In read-only mode openpyxl trusts that record
        # and yields the header alone, so three exports in a row were imported as
        # "zero observations" and written up as an export-settings problem. They
        # held 590, 736 and 775 rows. Recompute the extent instead of believing
        # the file, and keep read-only for the memory behaviour.
        worksheet.reset_dimensions()
        base = safe_sheet_name(worksheet.title)
        filename = f"{base}.tsv"
        if filename in used_names:
            filename = f"{index}-{base}.tsv"
        used_names.add(filename)
        target = output_dir / filename
        rows = 0
        columns = 0
        with target.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.writer(stream, dialect="excel-tab", lineterminator="\n")
            for row in worksheet.iter_rows(values_only=True):
                values = ["" if value is None else value for value in row]
                writer.writerow(values)
                rows += 1
                columns = max(columns, len(values))
        sheets.append({
            "sheet": worksheet.title,
            "file": filename,
            "rows_including_header": rows,
            "columns": columns,
        })
    workbook.close()

    manifest = {
        "source_id": args.source_id,
        "raw_filename": raw_target.name,
        "sha256": sha256(raw_target),
        "sheets": sheets,
    }
    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(manifest, ensure_ascii=False))


if __name__ == "__main__":
    main()
